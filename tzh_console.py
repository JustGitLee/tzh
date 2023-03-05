from openpyxl import load_workbook

file = 'tzh22.xlsx'
wb = load_workbook(file)
ws = wb.active


def trenya():
    columns = 'MNOL'
    mounth_columns = ['BC', 'BD', 'BE', 'BA']

    for i in range(len(columns)):
        while True:
            print(ws[columns[i] + '1'].value, end=' ')
            count = input('- ')
            if is_digit(count) == True:
                ws[columns[i] + str(checking_day)] = float(count)
                print(ws[mounth_columns[i] + '14'].value)
                break
    print(ws['BG12'].value)


def command():
    com = (input('Что записываем? "треня", "все" '))
    if com == 'треня':
        trenya()
    elif com == 'все':
        all()


def all():
    text = [3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 27, 30]
    dobro = [31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47]
    cell = 2
    for c in range(49):
        print(ws[1][cell].value, end=" ")
        print('(', ws[33][cell].value, ')', sep='', end=" ")

        if cell in text:
            data = input('- ')
            while True:
                if is_text(data) == False:
                    print(ws[1][cell].value, end=" ")
                    print('(', ws[33][cell].value, ')', sep='', end=" ")
                    data = input('- ')
                elif is_text(data) == True:
                    if data == '0':
                        cell = cell + 1
                        break
                    ws[checking_day][cell].value = data
                    cell = cell + 1
                    break
        elif cell in dobro:
            data = input('- ')
            while True:
                if is_dobro(data) == False:
                    print(ws[1][cell].value, end=" ")
                    print('(', ws[33][cell].value, ')', sep='', end=" ")
                    data = input('- ')
                elif is_dobro(data) == True:
                    if data == '0':
                        cell = cell + 1
                        break
                    ws[checking_day][cell].value = float(data)
                    cell += 1
                    break

        else:
            data = input('- ')
            while True:
                if is_digit(data) == False:
                    print(ws[1][cell].value, end=" ")
                    print('(', ws[33][cell].value, ')', sep='', end=" ")
                    data = input('- ')
                elif is_digit(data) == True:
                    if data == '0':
                        cell = cell + 1
                        break
                    ws[checking_day][cell].value = float(data)
                    cell += 1
                    break


def is_text(check):
    if check == '0' or check == 'х':
        return True
    else:
        print('Тут нужен "0" или "х"')
        return False

def is_dobro(check):
    if check == '0' or check == '1':
        return True
    else:
        print('Тут нужен "0" или "1"')
        return False

def is_digit(check):
    if check.isalpha():
        print('Число! Не текст.')
        return False
    else:
        return True


while True:
    day = input('День - ')
    if is_digit(day) == True:
        checking_day = int(day) + 1
        break

command()

wb.save(file)
wb.close()
