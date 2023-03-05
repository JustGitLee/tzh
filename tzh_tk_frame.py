from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook


file = 'tzh22.xlsx'
wb = load_workbook(file)
ws = wb.active
cell = 2

def get_data(event):
    global cell
    value = data.get()
    if value == '0':
        ent_day.delete(0, END)
        messagebox.showerror('!!!!!!', 'В следующиий раз жми Пропуск!')
        skip()
    elif value.isalpha():
        ent_day.delete(0, END)
        messagebox.showerror('!!!!!!', 'Введи число!')
    elif ',' in value:
        ent_day.delete(0, END)
        messagebox.showerror('!!!!!!', 'Убери ","!')
    else:
        ws[checking_day][cell].value = float(value)
        cell = cell + 1
        ent_all.delete(0, END)
        lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
        if ws[33][cell].value == 'х':
            frame_all.place_forget()
            frame_text.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
            lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
            lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
        else:
            frame_text.place_forget()
            frame_all.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
            lab_all1.configure(text=ws[1][cell].value, width=15, bd=2)
            lab_all2.configure(text=ws[33][cell].value, width=15, bd=2)

def terreble():
    global cell
    value = 1
    ws[checking_day][cell].value = int(value)
    cell = cell + 1
    lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)

def bad():
    global cell
    value = 2
    ws[checking_day][cell].value = int(value)
    cell = cell + 1
    lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)
    if cell == 51:
        wb.save(file)
        wb.close()
        root.quit()

def nothing():
    global cell
    value = 3
    ws[checking_day][cell].value = int(value)
    cell = cell + 1
    lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)
    if cell == 51:
        wb.save(file)
        wb.close()
        root.quit()

def good():
    global cell
    value = 4
    ws[checking_day][cell].value = int(value)
    cell = cell + 1
    lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)
    if cell == 51:
        wb.save(file)
        wb.close()
        root.quit()

def amazing():
    global cell
    value = 5
    ws[checking_day][cell].value = int(value)
    cell = cell + 1
    lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)
    if cell == 51:
        wb.save(file)
        wb.close()
        root.quit()

def done1():
    global cell
    value = 1
    ws[checking_day][cell].value = int(value)
    cell = cell + 1
    lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
    if ws[33][cell].value == 'х':
        frame_all.place_forget()
        frame_text.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
        but_text_done.configure(text='Есть', width=15, bd=2, bg='#26FF00', command=done1)
    elif ws[33][cell].value == '0-5':
        frame_text.place_forget()
        frame_evaluation.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)

def skip():
    global cell
    cell = cell + 1
    lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
    ent_day.delete(0, END)
    if cell == 11:
        cell = 15
    if ws[33][cell].value == 1:
        frame_all.place_forget()
        frame_text.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
        but_text_done.configure(text='Есть', width=15, bd=2, bg='#26FF00', command=done1)
    elif ws[33][cell].value == 'х':
        frame_all.place_forget()
        frame_text.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
    elif ws[33][cell].value == '0-5':
        frame_text.place_forget()
        frame_evaluation.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_evaluation1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_evaluation2.configure(text=ws[33][cell].value, width=15, bd=2)
    else:
        frame_text.place_forget()
        frame_all.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_all1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_all2.configure(text=ws[33][cell].value, width=15, bd=2)

def done():
    global cell
    value = 'х'
    ws[checking_day][cell].value = value
    cell = cell + 1
    lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
    if cell == 11:
        cell = 15
    if ws[33][cell].value == 1:
        frame_all.place_forget()
        frame_text.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
        but_text_done.configure(text='Есть', width=15, bd=2, bg='#26FF00', command=done1)
    elif ws[33][cell].value == 'х':
        frame_all.place_forget()
        frame_text.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_text1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_text2.configure(text=ws[33][cell].value, width=15, bd=2)
    else:
        frame_text.place_forget()
        frame_all.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
        lab_all1.configure(text=ws[1][cell].value, width=15, bd=2)
        lab_all2.configure(text=ws[33][cell].value, width=15, bd=2)

def get_day(event):
    day = data.get()
    global checking_day
    if day.isalpha():
        ent_day.delete(0, END)
        messagebox.showerror('!!!!!!', 'Введи число!')
    else:
        checking_day = int(day) + 1
        ent_day.delete(0, END)
        frame_day.place_forget()
        frame_menu.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)



def get_workout(event):
    global cell
    value = data.get()
    if ',' in value:
        ent_day.delete(0, END)
        messagebox.showerror('!!!!!!', 'Убери ","!')
    ws[checking_day][cell].value = float(value)
    ent_day.delete(0, END)
    cell = cell + 1
    lab_workout1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_workout2.configure(text=ws[33][cell].value, width=15, bd=2)
    if cell == 15:
        wb.save(file)
        wb.close()
        root.quit()

def rest_all():
    frame_menu.place_forget()
    frame_all.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)

def workout():
    frame_menu.place_forget()
    frame_workout.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
    global cell
    cell = 11
    lab_workout1.configure(text=ws[1][cell].value, width=15, bd=2)
    lab_workout2.configure(text=ws[33][cell].value, width=15, bd=2)


root = Tk()

data = StringVar()

root.title('Таблица Жизни')
root.geometry('400x100+650+300')
root.resizable(width=False, height=False)

frame_workout = Frame(root)
ent_workout = Entry(frame_workout, textvariable=data)
ent_workout.grid(column=1, row=0)
lab_workout1 = Label(frame_workout, text=ws[1][cell].value)
lab_workout1.grid(column=0, row=0)
lab_workout2 = Label(frame_workout, text=ws[33][cell].value)
lab_workout2.grid(column=0, row=1)

frame_menu = Frame(root)
but_menu_workout = Button(frame_menu, text='Тренировка', command=workout)
but_menu_workout.pack()
but_menu_rest_all = Button(frame_menu, text='Всё', command=rest_all)
but_menu_rest_all.pack()

frame_evaluation = Frame(root)
lab_evaluation1 = Label(frame_evaluation, text=ws[1][cell].value)
lab_evaluation1.grid(column=0, row=0)
lab_evaluation2 = Label(frame_evaluation, text=ws[33][cell].value)
lab_evaluation2.grid(column=0, row=1)
but_evaluation1 = Button(frame_evaluation, text='Ужасно', command=terreble)
but_evaluation1.grid(column=2, row=0)
but_evaluation2 = Button(frame_evaluation, text='Плохо', command=bad)
but_evaluation2.grid(column=3, row=0)
but_evaluation3 = Button(frame_evaluation, text='Никак', command=nothing)
but_evaluation3.grid(column=2, row=1)
but_evaluation4 = Button(frame_evaluation, text='Хорошо', command=good)
but_evaluation4.grid(column=2, row=2)
but_evaluation5 = Button(frame_evaluation, text='Замечательно', command=amazing)
but_evaluation5.grid(column=3, row=2)

frame_all = Frame(root)
lab_all1 = Label(frame_all, text=ws[1][cell].value)
lab_all1.grid(column=0, row=0)
lab_all2 = Label(frame_all, text=ws[33][cell].value)
lab_all2.grid(column=0, row=1)
ent_all = Entry(frame_all, textvariable=data)
ent_all.grid(column=1, row=0)
but_all_skip = Button(frame_all, text='Пропуск', command=skip)
but_all_skip.grid(column=2, row=0)

frame_text = Frame(root)
lab_text1 = Label(frame_text, text=ws[1][cell].value)
lab_text1.grid(column=0, row=0)
lab_text2 = Label(frame_text, text=ws[33][cell].value)
lab_text2.grid(column=0, row=1)
but_text_skip = Button(frame_text, text='Пропуск', command=skip)
but_text_skip.grid(column=2, row=0)
but_text_done = Button(frame_text, text='Есть', command=done)
but_text_done.grid(column=1, row=0)

frame_day = Frame(root)
frame_day.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.8)
lab_day = Label(frame_day, text='День')
lab_day.grid(column=0, row=0)
ent_day = Entry(frame_day, textvariable=data)
ent_day.grid(column=1, row=0)

ent_workout.bind('<Return>', get_workout)
ent_all.bind('<Return>', get_data)
ent_day.bind('<Return>', get_day)

root.mainloop()
