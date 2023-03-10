from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
import tkinter as tk

file = 'tzh23.xlsx'
wb = load_workbook(file)
ws = wb.active
cell = 2
xl_row_day = 0
last_value_cell = ws[33][2].value
current_value_cell = ws[33][2].value

def checking_cell():
    if not ws[33][cell].value and not ws[1][cell].value:
        save_close()

def back():
    global cell
    global current_frame
    cell -= 1
    checking_metric()
    current_frame.change_labels()

def change_frames(frame1, frame2):
    global current_frame
    frame1.hide()
    frame2.show()
    current_frame = frame2

def checking_metric():
    global current_frame
    global current_value_cell
    global last_value_cell
    last_value_cell = current_value_cell
    current_value_cell = ws[33][cell].value

    if current_value_cell == 'х' and last_value_cell != 'х':
        change_frames(current_frame, frame_mian_x)

    elif current_value_cell == 1 and last_value_cell != 1:
        change_frames(current_frame, frame_mian_dobro)

    elif current_value_cell == '0-5' and last_value_cell != '0-5':
        change_frames(current_frame, frame_mian_evaluation)

    elif (last_value_cell == 'х' and current_value_cell != 'х') or \
            (last_value_cell == '0-5' and current_value_cell != '0-5') or \
            (last_value_cell == 1 and current_value_cell != 1):
        change_frames(current_frame, frame_mian)

def save_close():
    wb.save(file)
    wb.close()
    root.quit()

def checking_number(number):
    global current_frame
    current_frame.clear_entry()
    if ',' in number:
        messagebox.showerror('!!!!!!', 'Убери ","!')
        return False
    elif number.isalpha() or (current_frame == frame_day and not number.isdigit()) or not number:
        messagebox.showerror('!!!!!!', 'Введи число!')
        return False
    elif current_frame == frame_day and number not in [str(i) for i in range(1, 32)]:
        messagebox.showerror('!!!!!!', 'Число месяца от 1 до 31!')
        return False
    return True

def skip():
    global cell
    global current_frame
    if ws[33][cell].value:
        ws[xl_row_day][cell].value = ''
    cell += 1
    checking_metric()
    current_frame.change_labels()

def done_x():
    global cell
    global current_frame
    value = 'х'
    ws[xl_row_day][cell].value = value
    cell += 1
    checking_metric()
    current_frame.change_labels()

def done_1():
    global cell
    global current_frame
    ws[xl_row_day][cell].value = 1
    cell += 1
    checking_metric()
    current_frame.change_labels()

def get_data(event):
    global cell
    global current_frame
    value = data.get()
    if checking_number(value):
        ws[xl_row_day][cell].value = float(value)
        cell += 1
        current_frame.clear_entry()
        checking_metric()
        current_frame.change_labels()

def get_day(event):
    day = data.get()
    if checking_number(day):
        global xl_row_day
        global current_frame
        xl_row_day = int(day) + 1
        change_frames(frame_day, frame_mian)

def evaluation(num):
    global cell
    global current_frame
    ws[xl_row_day][cell].value = num
    cell += 1
    checking_metric()
    current_frame.change_labels()
    checking_cell()

class Frame:
    def __init__(self, root):
        self.root = root  # атрибут класса
        self.frame = tk.Frame(root)  # атрибут класса

    def show(self):
        self.frame.pack()

    def hide(self):
        self.frame.pack_forget()

    def add_entry(self, textvariable, sequence, func):
        self.ent = tk.Entry(self.frame, textvariable=textvariable, width=15, bd=2)
        self.ent.grid(column=1, row=0)
        self.ent.bind(sequence, func)

    def clear_entry(self):
        self.ent.delete(0, END)

    def add_label_parameter(self, text, column, row):
        self.lab_parameter = tk.Label(self.frame, text=text, width=15, bd=2)
        self.lab_parameter.grid(column=column, row=row)

    def add_label_metric(self, text, column, row):
        self.lab_metric = tk.Label(self.frame, text=text, width=15, bd=2)
        self.lab_metric.grid(column=column, row=row)

    def change_labels(self):
        self.lab_parameter.configure(text=ws[1][cell].value, width=15, bd=2)
        self.lab_metric.configure(text=ws[33][cell].value, width=15, bd=2)

    def add_button(self, text, command, column, row):
        self.but = tk.Button(self.frame, text=text, command=command, width=15, bd=2)
        self.but.grid(column=column, row=row)

root = tk.Tk()
root.title('Таблица Жизни')
root.geometry('400x110+650+300')
root.resizable(width=False, height=False)

data = tk.StringVar()

frame_mian_evaluation = Frame(root)
frame_mian_evaluation.add_label_parameter(ws[1][cell].value, column=0, row=0)
frame_mian_evaluation.add_label_metric(ws[33][cell].value, column=0, row=1)
frame_mian_evaluation.add_button('Ужасно', lambda: evaluation(1), column=3, row=0)
frame_mian_evaluation.add_button('Плохо', lambda: evaluation(2), column=2, row=0)
frame_mian_evaluation.add_button('Никак', lambda: evaluation(3), column=2, row=1)
frame_mian_evaluation.add_button('Хорошо', lambda: evaluation(4), column=2, row=2)
frame_mian_evaluation.add_button('Замечательно', lambda: evaluation(5), column=3, row=2)
frame_mian_evaluation.add_button('назад', back, column=2, row=4)

frame_mian_dobro = Frame(root)
frame_mian_dobro.add_label_parameter(ws[1][cell].value, column=0, row=0)
frame_mian_dobro.add_label_metric(ws[33][cell].value, column=0, row=1)
frame_mian_dobro.add_button('Пропуск', skip, column=2, row=0)
frame_mian_dobro.add_button('Есть', done_1, column=1, row=0)
frame_mian_dobro.add_button('назад', back, column=2, row=4)

frame_mian = Frame(root)
frame_mian.add_label_parameter(ws[1][cell].value, column=0, row=0)
frame_mian.add_label_metric(ws[33][cell].value, column=0, row=1)
frame_mian.add_entry(data, '<Return>', get_data)
frame_mian.add_button('Пропуск', skip, column=2, row=0)
frame_mian.add_button('назад', back, column=2, row=4)

frame_mian_x = Frame(root)
frame_mian_x.add_label_parameter(ws[1][cell].value, column=0, row=0)
frame_mian_x.add_label_metric(ws[33][cell].value, column=0, row=1)
frame_mian_x.add_button('Пропуск', skip, column=2, row=0)
frame_mian_x.add_button('Есть', done_x, column=1, row=0)
frame_mian_x.add_button('назад', back, column=2, row=4)

frame_day = Frame(root)
frame_day.add_label_parameter('День', column=0, row=0)
frame_day.add_entry(data, '<Return>', get_day)

frame_day.show()
current_frame = frame_day

root.mainloop()
