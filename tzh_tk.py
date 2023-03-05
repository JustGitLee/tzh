from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook


file = 'tzh22.xlsx'
wb = load_workbook(file)
ws = wb.active
cell = 2
text = [2, 3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 27, 30]
dobro = [31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47]

def skip():
    global cell
    cell = cell + 1
    ent.delete(0, END)
    lab1 = Label(root, text=ws[1][cell].value, width=20, bd=2)
    lab1.grid(column=0, row=0)
    lab2 = Label(root, text=ws[33][cell].value, width=20, bd=2)
    lab2.grid(column=0, row=1)
    if cell == 20:
        wb.save(file)
        wb.close()
        root.quit()

def done():
    global cell
    ws[checking_day][cell].value = 'х'
    cell = cell + 1
    lab1 = Label(root, text=ws[1][cell].value, width=20, bd=2)
    lab1.grid(column=0, row=0)
    lab2 = Label(root, text=ws[33][cell].value, width=20, bd=2)
    lab2.grid(column=0, row=1)
    if cell == 20:
        wb.save(file)
        wb.close()
        root.quit()

def getting(event):
    global cell
    value = data.get()

    if cell in text:
        but_done = Button(root, text='Сделал', width=10, bd=2, command=done)
        but_done.grid(column=1, row=1)
        lab1 = Label(root, text=ws[1][cell].value, width=20, bd=2)
        lab1.grid(column=0, row=0)
        lab2 = Label(root, text=ws[33][cell].value, width=20, bd=2)
        lab2.grid(column=0, row=1)

    else:
        ws[checking_day][cell].value = float(value)
        cell = cell + 1
        lab1 = Label(root, text=ws[1][cell].value, width=20, bd=2)
        lab1.grid(column=0, row=0)
        lab2 = Label(root, text=ws[33][cell].value, width=20, bd=2)
        lab2.grid(column=0, row=1)
        ent.delete(0, END)
        if cell == 20:
            wb.save(file)
            wb.close()
            lab_end = Label(root, text="ВСЁ!", width=20, bd=2)
            lab_end.grid(column=2, row=1)
            root.quit()

def get_data_for_all():
    but_skip = Button(root, text='Пропуск', width=10, bd=2, command=skip)
    but_skip.grid(column=2, row=0)
    lab1 = Label(root, text=ws[1][cell].value, width=20, bd=2)
    lab1.grid(column=0, row=0)
    lab2 = Label(root, text=ws[33][cell].value, width=20, bd=2)
    lab2.grid(column=0, row=1)
    ent.bind('<Return>', getting)


def get_day(event):
    day = ent.get()
    if day.isalpha():
        messagebox.showerror('!!!!!!', 'Введи число!')
    ent.delete(0, END)
    global checking_day
    checking_day = int(day) + 1
    ent.unbind('<Return>')
    lab.forget()
    get_data_for_all()


root = Tk()

data = StringVar()

root.title('Таблица Жизни')
root.geometry('400x100+650+300')
root.resizable(width=False, height=False)

lab = Label(root, text='День', width=20, bd=2)
lab.grid(column=0, row=0)

ent = Entry(root, width=20, bd=2, textvariable=data)
ent.grid(column=1, row=0)

ent.bind('<Return>', get_day)

root.mainloop()
